from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Count, Q
from django.utils import timezone
from api.models.data.student_payment import StudentPayment
from api.models.data.student import Student, ClassLevel
from api.serializers.data.student_payment import StudentPaymentSerializer
from api.views.data.base import DataRootViewSet
from api.utils.excel_export import export_to_excel
from decimal import Decimal
from datetime import datetime
import calendar


class StudentPaymentViewSet(DataRootViewSet):
    queryset = StudentPayment.objects.all().order_by('-payment_date')
    serializer_class = StudentPaymentSerializer
    filterset_fields = ['student', 'payment_status', 'payment_date']
    search_fields = [
        'student__full_name', 'student__registration_number',
        'reference_number', 'description'
    ]

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by student
        student = self.request.query_params.get('student')
        if student:
            queryset = queryset.filter(student_id=student)

        # Filter by payment_cycle
        payment_cycle = self.request.query_params.get('payment_cycle')
        if payment_cycle:
            queryset = queryset.filter(payment_cycle=payment_cycle)

        # Filter by status
        status = self.request.query_params.get('payment_status')
        if status:
            queryset = queryset.filter(payment_status=status)

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)

        return queryset

    def perform_create(self, serializer):
        """Create student payment - journal entry created automatically by signal"""
        serializer.save()

    def perform_update(self, serializer):
        """Update student payment"""
        # Note: Journal entries are not updated automatically to maintain audit trail
        serializer.save()

    @action(detail=False, methods=['get'])
    def daily_summary(self, request):
        """Get daily payment summary"""
        date = request.query_params.get('date', timezone.now().date().isoformat())

        summary = StudentPayment.objects.filter(
            payment_date=date
        ).aggregate(
            total_amount=Sum('amount'),
            count=Count('id')
        )

        return Response({
            'date': date,
            'total_amount': float(summary['total_amount'] or 0),
            'payment_count': summary['count'] or 0
        })

    @action(detail=False, methods=['get'])
    def monthly_summary(self, request):
        """Get monthly payment summary"""
        year = request.query_params.get('year', timezone.now().year)
        month = request.query_params.get('month')

        queryset = StudentPayment.objects.filter(payment_date__year=year)

        if month:
            queryset = queryset.filter(payment_date__month=month)

        summary = queryset.aggregate(
            total_amount=Sum('amount'),
            count=Count('id')
        )

        return Response({
            'year': year,
            'month': month,
            'total_amount': float(summary['total_amount'] or 0),
            'payment_count': summary['count'] or 0
        })

    @action(detail=True, methods=['post'])
    def mark_as_paid(self, request, pk=None):
        """Mark payment as completed"""
        payment = self.get_object()
        payment.payment_status = 'completed'
        payment.save()

        return Response({
            'message': 'Payment marked as completed',
            'payment_status': payment.payment_status
        })

    @action(detail=True, methods=['post'])
    def mark_as_refunded(self, request, pk=None):
        """Mark payment as refunded"""
        payment = self.get_object()
        payment.payment_status = 'refunded'
        payment.save()

        return Response({
            'message': 'Payment marked as refunded',
            'payment_status': payment.payment_status
        })

    @action(detail=False, methods=['get'])
    def financial_info(self, request):
        """Get financial info for a student based on month/year"""
        student_id = request.query_params.get('student')
        month = request.query_params.get('month', timezone.now().month)
        year = request.query_params.get('year', timezone.now().year)
        
        if not student_id:
            return Response({'error': 'student parameter is required'}, status=400)
        
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)
        
        # Ensure month is zero-padded for consistent querying
        month_str = str(month).zfill(2)
        year_str = str(year)
        
        # Get payments for the specified month - check both padded and non-padded formats
        monthly_payments = StudentPayment.objects.filter(
            student=student,
            period_year=year_str,
            payment_status='completed',
            payment_cycle='monthly'
        ).filter(
            Q(period_month=month_str) | Q(period_month=str(month))
        ).aggregate(total_paid=Sum('amount'))['total_paid'] or Decimal('0')
        
        # Calculate yearly payments if payment cycle is yearly
        yearly_payments = StudentPayment.objects.filter(
            student=student,
            period_year=year_str,
            payment_status='completed',
            payment_cycle='yearly'
        ).aggregate(total_paid=Sum('amount'))['total_paid'] or Decimal('0')
        
        # Determine fee and payments based on payment cycle
        if student.payment_cycle == 'yearly':
            total_fee = student.yearly_fee
            paid_amount = yearly_payments
        else:
            total_fee = student.monthly_fee
            paid_amount = monthly_payments
        
        remaining = total_fee - paid_amount
        
        return Response({
            'student_id': student.id,
            'student_name': student.full_name,
            'currency': student.currency,
            'payment_cycle': student.payment_cycle,
            'monthly_fee': float(student.monthly_fee),
            'yearly_fee': float(student.yearly_fee),
            'period': {
                'month': int(month),
                'year': int(year)
            },
            'total_amount': float(total_fee),
            'paid_amount': float(paid_amount),
            'remaining_amount': float(remaining),
            'is_paid': paid_amount >= total_fee,
            'payment_percentage': float((paid_amount / total_fee * 100) if total_fee > 0 else 0)
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export student payments to Excel with monthly breakdown"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        class_level_ids = request.query_params.getlist('class_levels')
        
        if not start_date or not end_date:
            return Response({'error': 'start_date and end_date are required'}, status=400)
        
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        
        # Generate list of months in range
        months = []
        current_year = start.year
        current_month = start.month
        while (current_year, current_month) <= (end.year, end.month):
            months.append((current_year, current_month))
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1
        
        # Filter students by class levels if provided
        students_qs = Student.objects.filter(status='active').select_related('class_level')
        if class_level_ids:
            students_qs = students_qs.filter(class_level_id__in=class_level_ids)
        
        # If multiple class levels selected, create multi-tab Excel
        if class_level_ids and len(class_level_ids) > 1:
            return self._export_multi_tab_excel(students_qs, months, class_level_ids)
        
        # Single class level or all - create single tab Excel
        return self._export_single_excel(students_qs, months)
    
    def _get_student_monthly_data(self, student, year, month):
        """Get payment data for a student for a specific month"""
        month_str = str(month).zfill(2)
        year_str = str(year)
        
        payments = StudentPayment.objects.filter(
            student=student,
            period_year=year_str,
            payment_status='completed'
        ).filter(
            Q(period_month=month_str) | Q(period_month=str(month))
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        if student.payment_cycle == 'yearly':
            expected_fee = student.yearly_fee / 12
        else:
            expected_fee = student.monthly_fee
        
        remaining = max(expected_fee - payments, Decimal('0'))
        
        return {
            'paid': float(payments),
            'expected': float(expected_fee),
            'remaining': float(remaining)
        }
    
    def _export_single_excel(self, students_qs, months):
        """Export all students to a single Excel sheet"""
        headers = ['No', 'Student Name', 'Reg. Number', 'Class', 'Payment Cycle']
        for year, month in months:
            month_name = calendar.month_abbr[month]
            headers.extend([f'{month_name} {year} Paid', f'{month_name} {year} Remaining'])
        headers.extend(['Total Paid', 'Total Remaining'])
        
        data = []
        for idx, student in enumerate(students_qs, 1):
            row = [
                idx,
                student.full_name,
                student.registration_number,
                student.class_level.name if student.class_level else '-',
                student.payment_cycle
            ]
            total_paid = 0
            total_remaining = 0
            for year, month in months:
                month_data = self._get_student_monthly_data(student, year, month)
                row.extend([month_data['paid'], month_data['remaining']])
                total_paid += month_data['paid']
                total_remaining += month_data['remaining']
            row.extend([total_paid, total_remaining])
            data.append(row)
        
        filename = f"student_payments_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return export_to_excel(
            data=data,
            headers=headers,
            filename=filename,
            sheet_name='Student Payments',
            title='Student Payments Report'
        )
    
    def _export_multi_tab_excel(self, students_qs, months, class_level_ids):
        """Export students grouped by class level to multi-tab Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from django.http import HttpResponse
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        
        class_levels = ClassLevel.objects.filter(id__in=class_level_ids).order_by('level')
        
        for class_level in class_levels:
            sheet_name = class_level.name[:31]
            worksheet = workbook.create_sheet(title=sheet_name)
            
            headers = ['No', 'Student Name', 'Reg. Number', 'Payment Cycle']
            for year, month in months:
                month_name = calendar.month_abbr[month]
                headers.extend([f'{month_name} {year} Paid', f'{month_name} {year} Remaining'])
            headers.extend(['Total Paid', 'Total Remaining'])
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            students = students_qs.filter(class_level=class_level)
            for row_idx, student in enumerate(students, 2):
                row = [
                    row_idx - 1,
                    student.full_name,
                    student.registration_number,
                    student.payment_cycle
                ]
                total_paid = 0
                total_remaining = 0
                for year, month in months:
                    month_data = self._get_student_monthly_data(student, year, month)
                    row.extend([month_data['paid'], month_data['remaining']])
                    total_paid += month_data['paid']
                    total_remaining += month_data['remaining']
                row.extend([total_paid, total_remaining])
                
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1]))
                for row_idx in range(2, students.count() + 2):
                    val = worksheet.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 20)
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"student_payments_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
