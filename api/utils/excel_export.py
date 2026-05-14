from django.http import HttpResponse
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_to_excel(data, headers, filename, sheet_name='Sheet1', title=None, metadata=None, finance_summary=None):
    """
    Export data to Excel file.
    
    Args:
        data: List of lists/tuples containing row data
        headers: List of column headers
        filename: Name of the file to download
        sheet_name: Name of the worksheet
    
    Returns:
        HttpResponse with Excel file
    """
    try:
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            import csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(data)
            
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename.replace(".xlsx", ".csv")}"'
            return response
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]
        
        current_row = 1
        
        # Add title if provided
        if title:
            title_cell = worksheet.cell(row=current_row, column=1, value=title)
            title_cell.font = Font(bold=True, size=16, color='1f2937')
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
        
        # Add metadata if provided
        if metadata:
            for key, value in metadata.items():
                meta_cell = worksheet.cell(row=current_row, column=1, value=f'{key}: {value}')
                meta_cell.font = Font(size=11, color='6b7280')
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                current_row += 1
            current_row += 1
        
        # Add generation date
        date_cell = worksheet.cell(row=current_row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
        date_cell.font = Font(size=10, color='9ca3af')
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1
        
        # Add financial summary if provided
        if finance_summary:
            current_row += 1
            summary_title = worksheet.cell(row=current_row, column=1, value='FINANCIAL SUMMARY')
            summary_title.font = Font(bold=True, size=14, color='1f2937')
            summary_title.alignment = Alignment(horizontal='center')
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1
            
            # Table headers
            header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            table_headers = ['Category', 'Total', 'Paid', 'Remaining']
            for col_idx, header in enumerate(table_headers, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Parse finance_summary and display in table format
            data_fill = PatternFill(start_color='eff6ff', end_color='eff6ff', fill_type='solid')
            current_category = None
            category_data = {}
            
            for key, value in finance_summary.items():
                if key and not key.startswith('  '):
                    if current_category and category_data:
                        # Write previous category
                        cat_cell = worksheet.cell(row=current_row, column=1, value=current_category)
                        cat_cell.font = Font(bold=True, size=10)
                        cat_cell.fill = data_fill
                        cat_cell.border = border
                        
                        for col_idx, val in enumerate([category_data.get('total', ''), category_data.get('paid', ''), category_data.get('remaining', '')], start=2):
                            val_cell = worksheet.cell(row=current_row, column=col_idx, value=val)
                            val_cell.font = Font(size=10)
                            val_cell.fill = data_fill
                            val_cell.border = border
                            val_cell.alignment = Alignment(horizontal='right')
                        current_row += 1
                        category_data = {}
                    
                    current_category = key.replace(' Currency', '').replace(' Summary', '')
                elif key.startswith('  '):
                    clean_key = key.strip().lower()
                    if 'total' in clean_key:
                        category_data['total'] = value
                    elif 'paid' in clean_key:
                        category_data['paid'] = value
                    elif 'remaining' in clean_key:
                        category_data['remaining'] = value
            
            # Write last category
            if current_category and category_data:
                cat_cell = worksheet.cell(row=current_row, column=1, value=current_category)
                cat_cell.font = Font(bold=True, size=10)
                cat_cell.fill = data_fill
                cat_cell.border = border
                
                for col_idx, val in enumerate([category_data.get('total', ''), category_data.get('paid', ''), category_data.get('remaining', '')], start=2):
                    val_cell = worksheet.cell(row=current_row, column=col_idx, value=val)
                    val_cell.font = Font(size=10)
                    val_cell.fill = data_fill
                    val_cell.border = border
                    val_cell.alignment = Alignment(horizontal='right')
                current_row += 1
            
            # Set column widths for summary table
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 18
            worksheet.column_dimensions['C'].width = 18
            worksheet.column_dimensions['D'].width = 18
        
        current_row += 1
        
        # Write headers with formatting
        header_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        header_row = current_row
        current_row += 1
        
        # Write data
        for row_data in data:
            for col_idx, cell_data in enumerate(row_data, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=cell_data)
                cell.border = border
            current_row += 1
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row_idx in range(header_row + 1, current_row):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(max_length + 3, 30)
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        from rest_framework.response import Response
        from rest_framework import status
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
